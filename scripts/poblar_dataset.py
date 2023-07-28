import pprint as pp
import random
import sys

import json
import openpyxl
import psycopg2
import requests
import threading

TOKEN = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ1c2VyX2lkIjoyMSwidXNlcm5hbWUiOiJBbmRyZXNMb25kb1x1MDBmMW8iLCJleHAiOjE2ODc4NjA3ODIsImVtYWlsIjoic29wb3J0ZS50ZWNuaWNvM0BjYXRlbXBvcmFsZXMuY29tIn0.yI4dekue8a-7QurTcR0sJaJFh2go_1swM6YBlBBLqHQ'
HEADERS = {
    'Content-Type': 'application/json',
    'Authorization' : 'JWT ' + TOKEN
}


# # URls LOCAL HOST
# URLSERVICIO = 'http://127.0.0.1:8000/api/v1/core/servicio/'
# URLMETODOSPAGOS = 'http://127.0.0.1:8000/api/v1/core/tipo-pagos/'
# URLASESOR = 'http://127.0.0.1:8000/api/v1/core/comercial/'
# URLEMPRESA = 'http://127.0.0.1:8000/api/v1/core/empresa/'
# URLAFILIACION = 'http://127.0.0.1:8000/api/v1/afiliaciones/'


# URLCIUDADES = 'http://127.0.0.1:8000/api/v1/core/ciudades'
# URLARL = 'http://127.0.0.1:8000/api/v1/core/arl/'
# URLEMPRESAS = 'http://127.0.0.1:8000/api/v1/core/empresa/'
# URLTIPOSCOTIZANTE = 'http://127.0.0.1:8000/api/v1/core/tipo-cotizante/'
# URLSUBTIPO = 'http://127.0.0.1:8000/api/v1/core/subtipo-cotizante/'
# URLCARGO = 'http://127.0.0.1:8000/api/v1/core/cargo/'
# URLAFP = 'http://127.0.0.1:8000/api/v1/core/afp/'

# # URLS PRODUCCION
URLSERVICIO = 'http://sistema-tit.catemporales.com:8001/api/v1/core/servicio/'
URLMETODOSPAGOS = 'http://sistema-tit.catemporales.com:8001/api/v1/core/tipo-pagos/'
URLASESOR = 'http://sistema-tit.catemporales.com:8001/api/v1/core/comercial/'
URLEMPRESA = 'http://sistema-tit.catemporales.com:8001/api/v1/core/empresa/'
URLAFILIACION = 'http://sistema-tit.catemporales.com:8001/api/v1/afiliaciones/'


URLCIUDADES = 'http://sistema-tit.catemporales.com:8001/api/v1/core/ciudades'
URLARL = 'http://sistema-tit.catemporales.com:8001/api/v1/core/arl/'
URLEMPRESAS = 'http://sistema-tit.catemporales.com:8001/api/v1/core/empresa/'
URLTIPOSCOTIZANTE = 'http://sistema-tit.catemporales.com:8001/api/v1/core/tipo-cotizante/'
URLSUBTIPO = 'http://sistema-tit.catemporales.com:8001api/v1/core/subtipo-cotizante/'
URLCARGO = 'http://sistema-tit.catemporales.com:8001/api/v1/core/cargo/'
URLAFP = 'http://sistema-tit.catemporales.com:8001/api/v1/core/afp/'


def construir_servicios(sheet, creado_por:str):
    objects = []
    for row in range(2, sheet.max_row + 1):
        objects.append(json.dumps({
            "nombre": sheet.cell(row=row, column=1).value,
            "value": sheet.cell(row=row, column=2).value,
            "activo": True,
            "creado_por": creado_por
        }))
    return objects

def construir_metodo_pago(sheet, creado_por:str):
    objects = []
    for row in range(2, sheet.max_row + 1):
        objects.append({
            "nombre": sheet.cell(row=row, column=1).value,
            "activo": True,
            "creado_por": creado_por
        })
    return objects


def normalize(s):
    replacements = (("á", "a"),("é", "e"),("í", "i"),("ó", "o"),("ú", "u"))
    for a, b in replacements:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    return s


def search_city(value:str, cities):
    for city in cities:
        if normalize(city["name"].lower()) == normalize(value.lower()):
            return city['uuid']


def construir_asesores(sheet, creado_por:str):
    objects = []
    cities = get_cities()
    for row in range(2, sheet.max_row + 1):
        if not sheet.cell(row=row, column=1).value:
            break
        print(sheet.cell(row=row, column=1))
        objects.append(json.dumps({
            "tipo_documento": sheet.cell(row=row, column=1).value,
            "documento": sheet.cell(row=row, column=2).value,
            "nombre" : sheet.cell(row=row, column=3).value,
            "telefono" : sheet.cell(row=row, column=4).value,
            "correo" : sheet.cell(row=row, column=5).value,
            "direccion" : sheet.cell(row=row, column=6).value,
            "ciudad" : search_city(sheet.cell(row=row, column=7).value, cities),
            "activo": True,
            "creado_por": creado_por
        }))
    return objects

def construir_empresas_planilla(sheet, creado_por:str):
    objects = []
    for row in range(2, sheet.max_row + 1):
        objects.append(json.dumps({
            "tipo_empresa": sheet.cell(row=row, column=1).value,
            "tipo_documento": sheet.cell(row=row, column=2).value,
            "documento":sheet.cell(row=row, column=3).value,
            "digito_verificacion":sheet.cell(row=row, column=4).value,
            "nombre":sheet.cell(row=row, column=5).value,
            "tipo_documento_rep_legal": sheet.cell(row=row, column=6).value,
            "documento_rep_legal": sheet.cell(row=row, column=7).value,
            "rep_legal": sheet.cell(row=row, column=8).value,
            "ley_1607": True if sheet.cell(row=row, column=13).value == "SI" else False,
            "telefono_pla": sheet.cell(row=row, column=9).value,
            "correo_pla": sheet.cell(row=row, column=10).value,
            "direccion_pla": sheet.cell(row=row, column=11).value,
            "telefono": None,
            "direccion": None,
            "correo": None,
            "nombre_emp": None,
            "arl" : sheet.cell(row=row, column=12).value,
            "activo": True,
            "creado_por": creado_por
        }))
    return objects

def construir_empresa_empleadora(sheet, creado_por:str):
    objects = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == None:
            break
        objects.append(json.dumps({
            "arl" : None,
            "tipo_empresa": sheet.cell(row=row, column=1).value,
            "tipo_documento": sheet.cell(row=row, column=2).value,
            "documento":sheet.cell(row=row, column=3).value,
            "digito_verificacion":sheet.cell(row=row, column=4).value,
            "nombre":sheet.cell(row=row, column=5).value,
            "tipo_documento_rep_legal": None,
            "documento_rep_legal": None,
            "rep_legal": None,
            "ley_1607": None,
            "telefono_pla": None,
            "correo_pla": None,
            "direccion_pla": None,
            "telefono": sheet.cell(row=row, column= 6).value,
            "correo": sheet.cell(row=row, column= 7).value,
            "direccion": sheet.cell(row=row, column= 8).value,
            "nombre_emp": sheet.cell(row=row, column=5).value,
            "activo": True,
            "creado_por": creado_por
        }))
    return objects


def date(fecha):
    if fecha == None:
        return "1000-01-01"
    split_fecha = fecha.split('T')
    return split_fecha[0]


def define_arl(data, tarifa):
    if data is None:
        return None, True, None
    if tarifa == '0.522':
        nivel = 'I'
    elif tarifa == '1.046':
        nivel = 'II'
    elif tarifa == '2.436':
        nivel = 'III'
    elif tarifa == '4.35':
        nivel = 'IV'
    elif tarifa == '6.96':
        nivel = 'V'
    else: 
        nivel = 'I'
    return data, False, nivel

def construir_afiliaciones(sheet, creado_por:str):
    wookbook = openpyxl.Workbook()
    sheet_new = wookbook.active

    for row in range(2, sheet.max_row + 1):
        print('Reading row %d' % row)

        arl_empresa, lleva_arl, nivel_arl = define_arl(sheet.cell(row=row,column=34).value, sheet.cell(row=row,column=36).value)
    
        object = json.dumps({
            "tipo_documento": sheet.cell(row=row, column=1).value,
            "documento": sheet.cell(row=row, column=2).value,
            "primer_nombre": sheet.cell(row=row, column=3).value,
            "segundo_nombre": sheet.cell(row=row, column=4).value,
            "primer_apellido": sheet.cell(row=row, column=5).value,
            "segundo_apellido": sheet.cell(row=row, column=6).value,
            "tipo_empresa_planilla": sheet.cell(row=row, column=7).value, # Tomar UUID de planilla
            "tipo_empresa_empleadora": sheet.cell(row=row, column=8).value, # Tomar UUID de Empleadora
            "fecha_nacimiento": sheet.cell(row=row, column=9).value,
            "genero": sheet.cell(row=row, column=11).value, # Adaptar a los datos de la plataforma
            "tipo_cotizante": sheet.cell(row=row,column=12).value, # Tomar UUID del tipo_cotinzante
            "subtipo_cotizante": sheet.cell(row=row,column=13).value, # Tomar UUID del subtipo_cotizante
            "ibc": sheet.cell(row=row,column=14).value,
            "cuota": sheet.cell(row=row,column=15).value,
            "cotizante_ext": False if sheet.cell(row=row,column=16).value == 'No' else True,
            "telefono": sheet.cell(row=row,column=18).value,
            "celular": sheet.cell(row=row,column=19).value,
            "direccion": sheet.cell(row=row,column=20).value,
            "municipio": sheet.cell(row=row,column=21).value,  # Tomar UUID DE LA CIUDAD,
            "correo": sheet.cell(row=row,column=22).value if sheet.cell(row=row,column=22).value != None else "correofalso@falso.com",
            "cargo": sheet.cell(row=row,column=23).value, # Tomar UUID DEL CARGO
            "fecha_ingreso": date(sheet.cell(row=row,column=24).value),
            "afp": sheet.cell(row=row,column=25).value if sheet.cell(row=row,column=25).value != None else None, # tomar UUID DE LA AFP
            "tarifa_afp": sheet.cell(row=row,column=26).value, # Adaptar a los datos de la plataforma
            "ccf": sheet.cell(row=row,column=27).value, # tomar UUID DE LA CCF
            "tarifa_ccf": sheet.cell(row=row,column=28).value,  # Adaptar a los datos de la plataforma
            "sena": sheet.cell(row=row,column=29).value, # revisar datos
            "icbf": sheet.cell(row=row,column=30).value, # revisar datos
            "ciudad_ccf": sheet.cell(row=row,column=31).value, # tomar uuid de la ciudad
            "eps": sheet.cell(row=row,column=32).value, # tomar uuid de la eps
            "tarifa_eps": sheet.cell(row=row,column=33).value, # Adaptar a los datos de la plataforma
            "arl_empresa": arl_empresa, # tomar uuid de la ARL
            "tarifa_arl": sheet.cell(row=row,column=36).value, # Adaptar a los datos de la plataforma
            "comercial": sheet.cell(row=row, column=37).value, # tomar uuid del comercial
            "mensajeria": False if sheet.cell(row=row, column=39).value == 'No' else True,
            "observaciones": sheet.cell(row=row, column=41).value,
            "activo": False if sheet.cell(row=row, column=43).value == 'Inactivo' else True, # Adaptar a los datos de la plataforma

            "servicio": None,
            "tipo_salario": "F", # Preguntar y esto que 
            "nivel_arl": nivel_arl, # Preguntar y esto que
            "lleva_arl": lleva_arl,  # depende de la ARL si existe es True si no existe es false
            "creado_por": creado_por,
            "facturado": False
        })
        
        result = post_row_on_db(URLAFILIACION, object)
        if not result:
            source_row = sheet[row]
            sheet_new.append([cell.value for cell in source_row])
            wookbook.save("error.xlsx")
            


def get_data_from_excel(path:str, option:int, creado_por:str):
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.active
    if option == 1:
        objects = construir_servicios(sheet, creado_por)
    elif option == 2:
        objects = construir_metodo_pago(sheet, creado_por)
    elif option == 3:
        objects = construir_asesores(sheet, creado_por)
    elif option == 4:
        objects = construir_empresas_planilla(sheet, creado_por)
    elif option == 5:
        objects = construir_empresa_empleadora(sheet, creado_por)
    elif option == 7:
        objects = []
        construir_afiliaciones(sheet, creado_por)
    wb.close()
    return objects

def post_results(url:str, datas):
    for index, data_object in enumerate(datas):
        print("Uploading -> ", index + 2)
        try:
            response =  requests.post(url, data=data_object, headers=HEADERS)
            if response.status_code != 201:
                print("Response Extranio", end=' ')
                print(response.content)
        except Exception as e:
            print(f'Error {e}')
            
def post_row_on_db(url:str, data_object: dict):
    try:
        response =  requests.post(url, data=data_object, headers=HEADERS)
        if response.status_code != 201:
            print("Bad Response", end=' ')
            print(response.content)
            return False
    except Exception as e:
        print(f'Error {e}')
        return False
    return True

def get_cities():
    try:
        response = requests.get(url=URLCIUDADES, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting cities')
    
def get_cargos():
    try:
        response = requests.get(url=URLCARGO, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting cargos')

def get_afp():
    try:
        response = requests.get(url=URLAFP, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting cargos')
    
def get_empresas():
    try:
        response = requests.get(url=URLEMPRESAS, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting empresa')

def get_tipo_cotizante():
    try:
        response = requests.get(url=URLTIPOSCOTIZANTE, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting tipos cotizante')

def get_subtipo_cotizante():
    try:
        response = requests.get(url=URLSUBTIPO, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting subtipos')

def get_arl():
    try:
        response = requests.get(url=URLARL, headers=HEADERS)
        return response.json()
    except Exception as e:
        raise Exception('Error getting arls')


if __name__ == '__main__':
    if len(sys.argv) == 1 or len(sys.argv) > 3:
        print("Usage: python poblar_dataset [option] [uuid-user]")
        print('Options:\n\t -f --afiliaciones "Subir Afiliaciones"')
        print('\t -s --subir "Subir datos del dataset"')
        sys.exit(0)

    creado_por = sys.argv[2]

    path_base = './data_set.py'
    path_servicios = path_base + '/servicios.xlsx'
    path_medias_pago = path_base + '/Metodo_pago.xlsx'
    path_asesor = path_base + '/Asesor_Comercial.xlsx'
    path_empresa_planilla = path_base + '/Empresa_planilla.xlsx'
    path_empresa_empleadora = path_base + '/Empresa_empleadora.xlsx'
    path_afiliaciones = path_base + '/afiliaciones1.xlsx'

    if sys.argv[1] in ['-f', '--afiliaciones']:
        print("subir afiliaciones")
        get_data_from_excel(path_afiliaciones, 7, creado_por)
    elif sys.argv[1] in ['-s', '--subir']:
        """ upload servicios """ 
        # print("Subiendo Servicios")
        # result = get_data_from_excel(path_servicios, 1, creado_por)
        # post_results(URLSERVICIO, result)
        """ Upload Metodo Pago """
        #print("Subiendo Metodo Pago")
        #result = get_data_from_excel(path_medias_pago, 2, creado_por)
        #post_results(URLMETODOSPAGOS, result)
        """ Upload Asesor """
        # print("Subiendo Asesor")
        # result = get_data_from_excel(path_asesor, 3, creado_por)
        # post_results(URLASESOR, result)
        """ Upload Empresa Planilla"""
        # print("Subiendo empresa planilla")
        # result = get_data_from_excel(path_empresa_planilla, 4, creado_por)
        # post_results(URLEMPRESA, result)
        """ Upload Empresa Empleadora"""
        print("Subiendo empresa Empleadora")
        result = get_data_from_excel(path_empresa_empleadora, 5, creado_por)
        post_results(URLEMPRESA, result)
    else:
        print("Usage: python poblar_dataset [option] [uuid-user]")
        print('Options:\n\t -f --afiliaciones "Subir Afiliaciones"')
        print('\t -s --subir "Subir datos del dataset"')
        sys.exit(0)        
